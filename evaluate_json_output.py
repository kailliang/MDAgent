import json
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 读取原始 JSON 文件
with open('results/basic_331samples_11.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws.append(["question_id", "label", "response", "input_tokens", "output_tokens", "total_tokens", "difficulty", "correct", "total_correct", "accuracy"])

for idx, sample in enumerate(data, start=2):
    question_id = sample.get('question_id', '')
    label = sample.get('label', '')
    # response 可能为dict或str
    resp = sample.get('response', '')
    if isinstance(resp, dict):
        resp_val = resp.get('0.0', '')
        match = re.search(r'Answer:\s*([A-E])', resp_val)
        response = match.group(1) if match else ''
    else:
        match = re.search(r'Answer:\s*([A-E])', str(resp))
        response = match.group(1) if match else ''
    token_usage = sample.get('token_usage', {})
    input_tokens = token_usage.get('input_tokens', '')
    output_tokens = token_usage.get('output_tokens', '')
    total_tokens = token_usage.get('total_tokens', '')
    difficulty = sample.get('difficulty', '')
    ws.append([question_id, label, response, input_tokens, output_tokens, total_tokens, difficulty, None, None, None])
    # H列公式: =IF(Bx=Cx,1,0)
    ws[f'H{idx}'] = f'=IF(B{idx}=C{idx},1,0)'

# I2公式: =SUM(H2:H332)
ws['I2'] = '=SUM(H2:H332)'
# J2公式: =I2/331
ws['J2'] = '=I2/331'

wb.save('cleaned_output/basic_331samples_11_clear.xlsx')
print('Excel导出完成：cleaned_output/basic_331samples_11_clear.xlsx')
# 英文注释：This script extracts the uppercase letter after "Answer:" from the response field and exports the results to an Excel file, with formulas for correct count and accuracy.
# 中文注释：本脚本提取 response 字段中 "Answer:" 后的大写字母，并导出为 Excel 文件，并自动添加答对题数和准确率的公式。 